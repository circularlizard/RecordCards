#!/usr/bin/env python3
"""
Generate Record Cards from XLSX - Creates printable record cards from participant XLSX data.

Reads an XLSX file with a two-row header (row 1: category names merged across columns,
row 2: individual field names), normalises the data into the same structure expected by
the existing PDF-generation functions, and produces the same output PDF format.

The 'Patrol Name' column in the XLSX maps to the 'Section' concept used throughout the
existing code, and is displayed as the participant's group label on each record card.
"""

import os
import sys
import argparse
import datetime

import openpyxl

# Allow importing shared PDF-generation functions from the sibling module
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_record_cards import (  # noqa: E402
    format_age,
    format_address,  # noqa: F401 – re-exported for completeness
    create_compact_record_card,  # noqa: F401
    generate_pdf,
)

INPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'input')
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'output')
INPUT_FILE = os.path.join(INPUT_DIR, 'member-details.xlsx')
_today = datetime.datetime.now().strftime("%Y-%m-%d")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'RecordCards_{_today}.pdf')


def parse_xlsx_header(ws):
    """Parse the two-row merged-cell header from the worksheet.

    Row 1 holds category names.  In openpyxl, merged cells return a value only
    on the first cell of the merged region; subsequent cells in the same region
    return None.  Carrying the last non-None value forward reproduces the full
    category for every column.

    Row 2 holds individual field names.

    Returns a list (one entry per column) where each element is:
      - "Field"              for columns with no category (e.g. First name)
      - "Category: Field"   for categorised columns (e.g. Primary Contact 1: Phone 1)
      - None                for completely empty columns
    """
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))

    col_names = []
    current_category = None

    for cat_val, field_val in zip(row1, row2):
        # A non-empty row-1 cell marks the start of a new category block
        if cat_val is not None and str(cat_val).strip():
            current_category = str(cat_val).strip()

        if field_val is not None and str(field_val).strip():
            field = str(field_val).strip()
            col_names.append(f"{current_category}: {field}" if current_category else field)
        else:
            col_names.append(None)

    return col_names


def calculate_age(dob):
    """Calculate age from a date-of-birth value returned by openpyxl.

    Accepts a datetime.date or datetime.datetime object (openpyxl returns
    datetime.datetime for date-formatted cells when data_only=True).

    Returns a string in 'years / months' format, e.g. '14 / 05', which is
    compatible with format_age() and the adult/youth split logic in generate_pdf().
    Returns '' for missing or unrecognised values.
    """
    if dob is None:
        return ''
    if isinstance(dob, datetime.datetime):
        dob = dob.date()
    if not isinstance(dob, datetime.date):
        return ''

    today = datetime.date.today()
    years = today.year - dob.year
    months = today.month - dob.month

    if months < 0:
        years -= 1
        months += 12

    # Adjust if the birthday hasn't occurred yet this month
    if today.day < dob.day:
        months -= 1
        if months < 0:
            years -= 1
            months += 12

    return f"{years} / {months:02d}"


def read_xlsx(file_path):
    """Read the XLSX file and return a list of normalised participant dicts.

    The returned dicts use exactly the same keys as those produced by
    read_and_clean_csv() in generate_record_cards.py so that generate_pdf()
    (and all helper functions it calls) work without modification.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    col_names = parse_xlsx_header(ws)

    data = []
    processed_participants = set()

    for row in ws.iter_rows(min_row=3, values_only=True):
        # Skip completely empty rows
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue

        # Build a raw dict preserving original Python types from openpyxl
        # (date cells stay as datetime objects, numeric cells as int/float)
        raw = {}
        for col_name, value in zip(col_names, row):
            if col_name is not None:
                raw[col_name] = value

        def get_str(key):
            """Return a clean string for a raw cell value, or '' if absent/date."""
            v = raw.get(key)
            if v is None:
                return ''
            # Date/datetime objects are not meant to be displayed as strings here
            if isinstance(v, (datetime.date, datetime.datetime)):
                return ''
            # Excel sometimes stores numbers without leading zeros (e.g. phone numbers)
            if isinstance(v, float) and v == int(v):
                return str(int(v))
            if isinstance(v, float):
                return str(v)
            return str(v).strip()

        first = get_str('First name')
        last = get_str('Last name')
        if not first and not last:
            continue

        participant = {
            'First Name':   first,
            'Last Name':    last,
            # 'Section' is the key used throughout generate_record_cards.py;
            # it is populated from the 'Patrol Name' column in the XLSX.
            'Section':      get_str('Patrol Name'),
            'Age at Start': calculate_age(raw.get('Date of birth')),
            # The XLSX has no 'Swimmer' equivalent; set to empty string so the
            # existing create_compact_record_card() receives a valid key.
            'Swimmer':      '',
        }

        # Contact blocks — Primary Contact 1, Primary Contact 2, Emergency Contact
        for prefix in ('Primary Contact 1', 'Primary Contact 2', 'Emergency Contact'):
            for field in ('First Name', 'Last Name', 'Address 1', 'Address 2',
                          'Address 3', 'Address 4', 'Postcode', 'Email 1', 'Email 2',
                          'Phone 1', 'Phone 2'):
                participant[f'{prefix}: {field}'] = get_str(f'{prefix}: {field}')

        # Medical / essential information
        for field in ('Medical details', 'Allergies', 'Dietary requirements',
                      'Tetanus (year of last jab)', 'Other useful information'):
            participant[field] = get_str(f'Essential Information: {field}')

        # Member's own contact details (columns AW–BE in the XLSX)
        for field in ('Address 1', 'Address 2', 'Address 3', 'Address 4', 'Postcode',
                      'Email 1', 'Email 2', 'Phone 1', 'Phone 2'):
            participant[f'Member: {field}'] = get_str(f'Member: {field}')

        # Additional information (columns BL–BM in the XLSX)
        participant['Additional Information: Additional Needs'] = get_str(
            'Additional Information: Additional Needs'
        )
        participant['Additional Information: SEEE Expeditions Extra notes from parents'] = get_str(
            'Additional Information: SEEE Expeditions Extra notes from parents'
        )

        # Deduplication
        pid = (f"{participant['First Name']}_"
               f"{participant['Last Name']}_"
               f"{participant['Section']}")
        if pid in processed_participants:
            continue
        processed_participants.add(pid)

        # Clean phone numbers (remove underscores used for visual formatting in Excel)
        for key in participant:
            if 'Phone' in key and participant[key]:
                participant[key] = participant[key].replace('_', '')

        # Handle 'as above' emergency contact — copy Primary Contact 1 details
        if participant['Emergency Contact: First Name'].strip().lower() == 'as above':
            for field in ('First Name', 'Last Name', 'Address 1', 'Address 2',
                          'Address 3', 'Address 4', 'Postcode', 'Phone 1', 'Phone 2'):
                participant[f'Emergency Contact: {field}'] = participant[f'Primary Contact 1: {field}']

        participant['Age Formatted'] = format_age(participant['Age at Start'])

        data.append(participant)

    wb.close()

    # Sort by patrol name (Section) then last name — mirrors the CSV script behaviour
    data.sort(key=lambda x: (x['Section'], x['Last Name']))
    return data


def main():
    parser = argparse.ArgumentParser(
        description='Generate record cards from participant XLSX data'
    )
    parser.add_argument('--input', help='Path to input XLSX file', default=INPUT_FILE)
    parser.add_argument('--output', help='Path to output PDF file', default=OUTPUT_FILE)
    parser.add_argument('--records-per-page', type=int, default=8,
                        help='Number of records per page')
    args = parser.parse_args()

    os.makedirs(os.path.dirname(args.output), exist_ok=True)

    print(f"Reading data from: {args.input}")
    data = read_xlsx(args.input)
    print(f"Found {len(data)} participants")

    generate_pdf(data, args.output, args.records_per_page)


if __name__ == '__main__':
    main()
